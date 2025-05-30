from openpyxl import load_workbook
wb = load_workbook("book1.xlsx", read_only=True, data_only=True)
print(wb.sheetnames)
